
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from datetime import datetime
import openpyxl 
import re
import semver

####DO NOT CHANGE UNLESS THE FORTINET DOC WEBSITE CHANGES
URLperm='https://docs.fortinet.com'
##### Please give the full directory path of the (SSOE2 Software Inventory - Network (Tab 9).xlsx) file. Remember to add the extension e.g. .xlsx .xls:
### use // instead of / for the file path
path="C:\\Users\\P1350392\\OneDrive - NCS Pte Ltd\\work proj files\\SSOE2 Software Inventory - Network (Tab 9) - Copy.xlsx"


#####################################################################################################
#################### Function to get relevant data
def get_data(URL,integral):

            
 
  
#### list we are going to use to store values
  mainpage=[]
  subpage=[]
  dates=[]
  dict={}
  final_date=["1000-10-1"]

  page = requests.get(URL)
  soup = BeautifulSoup(page.text,'lxml')
  ## getting all links from webpage
  for z in soup.findAll('a',{'class':"version-item-external"}):
      dir=z.get('href')
      
  ## searching only for release notes
      if re.search(integral,dir):
        mainpage.append(URLperm+dir)

#loping through mini release links 
  for i in mainpage:
    page2=requests.get(i)
    soup2=BeautifulSoup(page2.text,'lxml')

  ### getting all links from release notes page
    for a in soup2.findAll('a',{'class':"toc"}):

  ## wanting only change log page   
      dir2=a.get('href')
      if (re.search('change-log$',dir2)) and (URLperm+dir2 not in subpage) :
        subpage.append(URLperm+dir2)

  #part of the code to get the dates
  for x in subpage:
    page3=requests.get(x)
    soup3=BeautifulSoup(page3.text,'lxml')
    table=soup3.find('table',{'class':'TableStyle-FortinetTable'})
  
    for date in table.find('td'):
      if date.text != '\n':
        i = re.sub(r'[\t\n ]+', ' ', date.text).strip()
        dates.append(i)

    
  # creating a new dictionary key value pair with the key being the link and the value being the date
  for key1 in subpage:
    for key2 in dates:
      dict[key1]=key2
      dates.remove(key2)
      break

  # finding the latest date the latest date
  for values in dict.values():

    try:
      from_value = datetime.strptime(values, '%Y-%m-%d')
      latest_date= datetime.strptime(final_date[0],'%Y-%m-%d')
    
      if from_value > latest_date :
        final_date.clear()
        final_date.append(values)
    except ValueError as message:
      print('A value error is raised because :', message)

  final_date=final_date[0]
  final_link = list(filter(lambda x: dict[x] == final_date, dict))[0]

#getting the version value
  page4=requests.get(final_link)
  soup4=BeautifulSoup(page4.text,'lxml')

  latest_version=soup4.find('span',{'class':'current-version'}).text
  #after getting the value we return the data to append later
  return latest_version,final_date,final_link


 ######################## getting the latest major market version###################################
def latest_major(URL,integral):
 
  major_versions=[]
  major_versions_url=[]
  latest_major="0.0.0"
  major_mainpage=[]
  major_subpage=[]
  major_dates=[]
  major_dict={}
  date1=['1000-10-1']

##### major version number
  page = requests.get(URL)
  soup = BeautifulSoup(page.text,'lxml')

  for i in soup.findAll('a',{'class':"version-family-item"}):
      link=i.get('href')
      i = re.sub(r'[\t\n ]+', ' ', i.text).strip()
      
      if i+".0" not in major_versions:
        major_versions.append(i+".0")
        major_versions_url.append(URLperm+link)

  #comparing the major version and getting the latest
  for i in major_versions:
    if semver.compare(i,latest_major)==1:
     latest_major=i

  
# getting the exact link for the mainpage 
  ver=latest_major.replace('.0','')
  for c in major_versions_url:
    if c.endswith(ver):
      
      page5 = requests.get(c)
      soup5 = BeautifulSoup(page5.text,'lxml')
      for z in soup5.findAll('a',{'class':"version-item-external"}):
          dir5=z.get('href')
          if dir5.endswith(integral):
            major_mainpage.append(URLperm+dir5)
  
  
  for i in major_mainpage:
    page2=requests.get(i)
    soup2=BeautifulSoup(page2.text,'lxml')

  ### getting all links from release notes page
    for a in soup2.findAll('a',{'class':"toc"}):

  ## wanting only change log page   
      dir2=a.get('href')
      if (re.search('change-log$',dir2)) and (URLperm+dir2 not in major_subpage) :
        major_subpage.append(URLperm+dir2)
  
  #part of the code to get the dates

  for x in major_subpage:
    page3=requests.get(x)
    soup3=BeautifulSoup(page3.text,'lxml')
    table=soup3.find('table',{'class':'TableStyle-FortinetTable'})
  
    for date in table.find('td'):
      if date.text != '\n':
        i = re.sub(r'[\t\n ]+', ' ', date.text).strip()
        major_dates.append(i)

    
  # creating a new dictionary key value pair with the key being the link and the value being the date
  for key1 in major_subpage:
    for key2 in major_dates:
      major_dict[key1]=key2
      major_dates.remove(key2)
      break
  
 # finding the latest date the latest date
  
  for values in major_dict.values():
    try:
      
      from_value = datetime.strptime(values,'%Y-%m-%d')
      date= datetime.strptime(date1[0],'%Y-%m-%d')
      if from_value > date :
        date1.clear()
        date1.append(values)
    except ValueError as message:
      print('A value error is raised because :', message)
  #after getting the value we return the data to append later
  
  major_date=date1[0]
  
  major_link = list(filter(lambda x: major_dict[x] == major_date, major_dict))[0]

  page4=requests.get(major_link)
  soup4=BeautifulSoup(page4.text,'lxml')

  major_version=soup4.find('span',{'class':'current-version'}).text
  
  return major_version,major_date


  
  
#######################################################################################################
############## function for fortigate
def fortigate(latest_version,final_date,final_link,major_ver, major_date):
  try  : 
    wb_obj = openpyxl.load_workbook(path) 
  except ValueError as message:
    print('A value error is raised because :', message)

  try:
    sheet= wb_obj.active 
    #slotting in the latest current version avail
    cell = sheet['K10']
    cell.value= latest_version
    #slotting in current version latest release date
    cell2 =sheet['N10']
    cell2.value=final_date

    #slotting in the link into the file
    cell3=sheet['F10']
    cell3.value=final_link

    cell4=sheet['L10']
    cell4.value=major_ver

    cell5=sheet['M10']
    cell5.value=major_date

    #updating remarks 
    cellremarks=sheet['AA10']
    today = datetime.today()
    d2 = today.strftime("%Y-%m-%d")
    current_time = datetime.now()
    fmt_current_time=current_time.strftime("%H:%M:%S")
    cellremarks.value=d2+" "+ fmt_current_time + " (bot) : Updated by bot at this time"
    wb_obj.save(filename=path)
    print('fortigate field has been updated')
  except ValueError as message:
    print('A value error is raised because :', message)   
    print ('unable to append data to file') 

######################################################################################################
########## Function for FortiAnalyzer


def fortianalyzer(latest_version,final_date,final_link,major_ver, major_date):
  try  : 
    wb_obj = openpyxl.load_workbook(path) 
  except ValueError as message:
    print('A value error is raised because :', message)

  try:
    sheet= wb_obj.active 
    #slotting in the latest current verion avail
    cell = sheet['K12']
    cell.value= latest_version
    #slotting in current version latest release date
    cell2 =sheet['N12']
    cell2.value=final_date

    #slotting in the link into the file
    cell3=sheet['F12']
    cell3.value=final_link


    cell4=sheet['L12']
    cell4.value=major_ver

    cell5=sheet['M12']
    cell5.value=major_date
    #updating remarks 
    cellremarks=sheet['AA12']
    today = datetime.today()
    d2 = today.strftime("%Y-%m-%d")
    current_time = datetime.now()
    fmt_current_time=current_time.strftime("%H:%M:%S")
    cellremarks.value=d2+" "+ fmt_current_time + " (bot) : Updated by bot at this time"
    wb_obj.save(filename=path)
    print('fortianalyzer field has been updated')
  except ValueError as message:
    print('A value error is raised because :', message)   
    print ('unable to append data to file')

#################################################################################################################
########## Function for Fortimanager
def fortimanager(latest_version,final_date,final_link,major_ver, major_date):
  try  : 
    wb_obj = openpyxl.load_workbook(path) 
  except ValueError as message:
    print('A value error is raised because :', message)

  try:
    sheet= wb_obj.active 
    #slotting in the latest current verion avail
    cell = sheet['K13']
    cell.value= latest_version
    #slotting in current version latest release date
    cell2 =sheet['N13']
    cell2.value=final_date

    #slotting in the link into the file
    cell3=sheet['F13']
    cell3.value=final_link

    cell4=sheet['L13']
    cell4.value=major_ver

    cell5=sheet['M13']
    cell5.value=major_date
    #updating remarks 
    cellremarks=sheet['AA13']
    today = datetime.today()
    d2 = today.strftime("%Y-%m-%d")
    current_time = datetime.now()
    fmt_current_time=current_time.strftime("%H:%M:%S")
    cellremarks.value=d2+" "+ fmt_current_time + " (bot) : Updated by bot at this time"
    wb_obj.save(filename=path)
    print('Fortimanager field has been updated')
  except ValueError as message:
    print('A value error is raised because :', message)   
    print ('unable to append data to file')

########################################################################################################
#####RUNNING ALL THE FUNCTIONS HERE


URL="https://docs.fortinet.com/product/fortigate/7.0"
data =get_data(URL,'fortios-release-notes$')
latest_version=data[0]
final_date=data[1]
final_link=data[2]
major=latest_major(URL,'fortios-release-notes')
major_ver=major[0]
major_date=major[1]
fortigate(latest_version,final_date,final_link, major_ver, major_date)




URL1="https://docs.fortinet.com/product/fortianalyzer/7.0"
data2=get_data(URL1,'release-notes$')
latest_version2=data2[0]
final_date2=data2[1]
final_link2=data2[2]
major2=latest_major(URL1,'release-notes')
major_ver2=major2[0]
major_date2=major2[1]
fortianalyzer(latest_version2,final_date2,final_link2, major_ver2, major_date2)


URL2="https://docs.fortinet.com/product/fortimanager/7.0"
data3=get_data(URL2,'release-notes$')
latest_version3=data3[0]
final_date3=data3[1]
final_link3=data3[2]
major3=latest_major(URL2,'release-notes')
major_ver3=major2[0]
major_date3=major2[1]
fortimanager(latest_version3,final_date3,final_link3, major_ver3, major_date3)



   